## 1 pip install openpyxl
import openpyxl

def check_excel_cell_and_read_multiple(file_paths, cell, expected_value, sheet_names):
    matched_files_data = []

    for file_path in file_paths:
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)

        # Iterate through each worksheet
        for sheet_name in sheet_names:
            try:
                # Select the worksheet
                sheet = wb[sheet_name]

                # Get the value of the specified cell
                cell_value = sheet[cell].value

                # Check if the cell value matches the expected value
                if cell_value == expected_value:
                    # Read the worksheet data
                    data = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        data.append(row)

                    matched_files_data.append({
                        'file_path': file_path,
                        'sheet_name': sheet_name,
                        'data': data
                    })
            except KeyError:
                print(f"Worksheet '{sheet_name}' does not exist in the Excel file '{file_path}'.")
            except Exception as e:
                print(f"An error occurred while processing worksheet '{sheet_name}' in the Excel file '{file_path}': {str(e)}")

    return matched_files_data

# Example usage
file_paths = ["file1.xlsx", "file2.xlsx", "file3.xlsx"]
cell = "A1"
expected_value = "Hello"
sheet_names = ["Sheet1", "Sheet2"]

data = check_excel_cell_and_read_multiple(file_paths, cell, expected_value, sheet_names)
if data:
    for file_data in data:
        print(f"Data in worksheet '{file_data['sheet_name']}' of file '{file_data['file_path']}':")
        for row in file_data['data']:
            print(row)
else:
    print("No matches found in any of the specified files.")

# Description:
"""
In this script, the check_excel_cell_and_read_multiple function takes four parameters:

file_paths: A list of file paths to the Excel files.
cell: The cell reference (e.g., "A1", "B2", etc.) to check for a value match.
expected_value: The expected value that should match the value of the specified cell.
sheet_names: A list of worksheet names to check for a match.
The function iterates through each specified file and worksheet, checks the value of the specified cell, and compares it with the expected value. If there's a match, it reads the data from the worksheet and returns it as a list of dictionaries containing the file path, sheet name, and data.

To use the script:

Save the check_excel_cell_and_read_multiple.py file in the same directory where you have the Excel files you want to check.
Replace "file1.xlsx", "file2.xlsx", "file3.xlsx" with the actual file names or paths of your Excel files.
Specify the cell reference ("A1"), the expected value, and the worksheet names you want to check.
Run the script using a Python interpreter.
The script will check each specified file and worksheet for a match of the cell value with the expected value. If there's a match, it will read the data from the matched worksheet(s) and display it row by row, along with the file path and sheet name. If no matches are found in any of the specified files, it will display a message indicating that.

"""
## 2 
""" 
Certainly! Heres an example of a modular Python script that checks the value of a specific cell in an Excel file matches it with a variable value, and reads the worksheet data if there's a match:

Install the required libraries:
python
Copy code
pip install openpyxl
Create a modular script check_excel_cell_and_read.py:

Code: 
"""
import openpyxl

def check_excel_cell_and_read(file_path, cell, expected_value, sheet_names):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Iterate through each worksheet
    for sheet_name in sheet_names:
        try:
            # Select the worksheet
            sheet = wb[sheet_name]

            # Get the value of the specified cell
            cell_value = sheet[cell].value

            # Check if the cell value matches the expected value
            if cell_value == expected_value:
                # Read the worksheet data
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data.append(row)

                return data
        except KeyError:
            print(f"Worksheet '{sheet_name}' does not exist in the Excel file.")
        except Exception as e:
            print(f"An error occurred while processing worksheet '{sheet_name}': {str(e)}")

    return None

# Example usage
file_path = "example.xlsx"
cell = "A1"
expected_value = "Hello"
sheet_names = ["Sheet1", "Sheet2"]

data = check_excel_cell_and_read(file_path, cell, expected_value, sheet_names)
if data is not None:
    print(f"Data in matched worksheet(s):")
    for row in data:
        print(row)

# Description:
"""
In this script, the check_excel_cell_and_read function takes four parameters:

file_path: The file path to the Excel file.
cell: The cell reference (e.g., "A1", "B2", etc.) to check for a value match.
expected_value: The expected value that should match the value of the specified cell.
sheet_names: A list of worksheet names to check for a match.
The function iterates through each specified worksheet, checks the value of the specified cell, and compares it with the expected value. If there's a match, it reads the data from the worksheet and returns it as a list of rows.

To use the script:

Save the check_excel_cell_and_read.py file in the same directory where you have the Excel file you want to check.
Replace "example.xlsx" with the actual file name or path of your Excel file.
Specify the cell reference ("A1"), the expected value, and the worksheet names you want to check.
Run the script using a Python interpreter.
The script will check each specified worksheet for a match of the cell value with the expected value. If there's a match, it will read the data from the matched worksheet(s) and display it row by row.

"""



